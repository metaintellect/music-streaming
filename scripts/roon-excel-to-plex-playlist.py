#!/usr/bin/env python3
"""
Convert Roon Excel playlist exports to M3U files and optionally import them into Plex.

Examples:
    # Generate two playlist backups only
    python3 scripts/roon-excel-to-plex-playlist.py \
      --xlsx ~/Documents/exyu.xlsx ~/Documents/easy.xlsx \
      --output-dir /Volumes/xnas/playlists \
      --generate-only

    # Generate and import into Plex
    PLEX_TOKEN=... python3 scripts/roon-excel-to-plex-playlist.py \
      --xlsx ~/Documents/exyu.xlsx ~/Documents/easy.xlsx \
      --output-dir /Volumes/xnas/playlists \
      --plex-path-dir "\\\\192.168.100.83\\xnas\\playlists" \
      --plex-url http://WINDOWS_IP:32400 \
      --library-name Music \
      --import
"""

from __future__ import annotations

import argparse
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised only on machines missing openpyxl
    print("ERROR: openpyxl is required. Install it with: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(2)


DEFAULT_SHEET = "Tracks"
DEFAULT_PATH_COLUMN = "Path"


class PlaylistError(Exception):
    """User-facing error."""


@dataclass(frozen=True)
class PlaylistJob:
    xlsx: Path
    playlist_name: str
    output: Path
    plex_path: str | None


@dataclass(frozen=True)
class ParsedPlaylist:
    paths: list[str]
    tracks_read: int
    blank_paths: int


def parse_roon_excel(xlsx: Path, sheet_name: str = DEFAULT_SHEET, path_column: str = DEFAULT_PATH_COLUMN) -> ParsedPlaylist:
    if not xlsx.exists():
        raise PlaylistError(f"Excel file not found: {xlsx}")

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise PlaylistError(f"Sheet '{sheet_name}' not found in {xlsx}. Available sheets: {', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise PlaylistError(f"Sheet '{sheet_name}' is empty in {xlsx}") from None

    header_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
    if path_column not in header_map:
        raise PlaylistError(f"Column '{path_column}' not found in {xlsx}. Columns: {', '.join(header_map)}")

    path_index = header_map[path_column]
    paths: list[str] = []
    blank_paths = 0
    tracks_read = 0

    for row in rows:
        if not row or all(value is None for value in row):
            continue
        tracks_read += 1
        value = row[path_index] if path_index < len(row) else None
        if value is None or str(value).strip() == "":
            blank_paths += 1
            continue
        paths.append(str(value).strip())

    return ParsedPlaylist(paths=paths, tracks_read=tracks_read, blank_paths=blank_paths)


def to_relative_paths(paths: list[str], relative_root: str) -> list[str]:
    normalized_root = relative_root.rstrip("\\/")
    relative_paths: list[str] = []

    for path in paths:
        normalized_path = path.rstrip()
        if normalized_path.lower().startswith(normalized_root.lower() + "\\"):
            relative_paths.append(normalized_path[len(normalized_root) + 1 :])
        elif normalized_path.lower().startswith(normalized_root.lower() + "/"):
            relative_paths.append(normalized_path[len(normalized_root) + 1 :])
        else:
            raise PlaylistError(f"Path is not under relative root '{relative_root}': {path}")

    return relative_paths


def write_m3u(output: Path, paths: list[str], create_dirs: bool = False) -> None:
    parent = output.parent
    if not parent.exists():
        if create_dirs:
            try:
                parent.mkdir(parents=True, exist_ok=True)
            except OSError as exc:
                raise PlaylistError(f"Could not create output directory {parent}: {exc}") from exc
        else:
            raise PlaylistError(f"Output directory does not exist: {parent}. Pass --create-dirs to create it.")

    try:
        with output.open("w", encoding="utf-8", newline="\n") as file:
            for path in paths:
                file.write(f"{path}\n")
    except OSError as exc:
        raise PlaylistError(f"Could not write M3U file {output}: {exc}") from exc


def plex_get_xml(plex_url: str, endpoint: str, token: str, timeout: int) -> ET.Element:
    url = build_plex_url(plex_url, endpoint, {"X-Plex-Token": token})
    request = urllib.request.Request(url, headers={"Accept": "application/xml"})
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return ET.fromstring(response.read())
    except urllib.error.URLError as exc:
        raise PlaylistError(f"Plex server is unreachable at {plex_url}: {exc}") from exc
    except ET.ParseError as exc:
        raise PlaylistError(f"Plex returned invalid XML for {endpoint}: {exc}") from exc


def build_plex_url(plex_url: str, endpoint: str, query: dict[str, str | int]) -> str:
    base = plex_url.rstrip("/")
    path = endpoint if endpoint.startswith("/") else f"/{endpoint}"
    encoded_query = urllib.parse.urlencode(query)
    return f"{base}{path}?{encoded_query}"


def find_library_section_id(plex_url: str, token: str, library_name: str, timeout: int) -> tuple[str, str]:
    root = plex_get_xml(plex_url, "/library/sections", token, timeout)
    matches: list[tuple[str, str]] = []

    for directory in root.findall("Directory"):
        title = directory.attrib.get("title", "")
        key = directory.attrib.get("key", "")
        section_type = directory.attrib.get("type", "")
        if title.lower() == library_name.lower():
            matches.append((key, section_type))

    if not matches:
        available = ", ".join(
            directory.attrib.get("title", "<unnamed>") for directory in root.findall("Directory")
        )
        raise PlaylistError(f"No Plex library named '{library_name}' found. Available libraries: {available or 'none'}")

    section_id, section_type = matches[0]
    if not section_id:
        raise PlaylistError(f"Plex library '{library_name}' did not include a section key")

    return section_id, section_type


def upload_playlist(
    plex_url: str,
    token: str,
    section_id: str,
    plex_path: str,
    force: bool,
    timeout: int,
) -> tuple[int, str]:
    query: dict[str, str | int] = {
        "sectionID": section_id,
        "path": plex_path,
        "X-Plex-Token": token,
    }
    if force:
        query["force"] = 1

    url = build_plex_url(plex_url, "/playlists/upload", query)
    request = urllib.request.Request(url, method="POST", headers={"Accept": "application/xml"})
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            body = response.read().decode("utf-8", errors="replace")
            return response.status, body
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise PlaylistError(f"Plex playlist upload failed with HTTP {exc.code}: {body}") from exc
    except urllib.error.URLError as exc:
        raise PlaylistError(f"Plex server is unreachable at {plex_url}: {exc}") from exc


def resolve_token(args: argparse.Namespace) -> str | None:
    return args.plex_token or os.environ.get("PLEX_TOKEN")


def build_jobs(args: argparse.Namespace) -> list[PlaylistJob]:
    xlsx_paths = [Path(path).expanduser() for path in args.xlsx]

    if len(xlsx_paths) > 1 and args.playlist_name:
        raise PlaylistError("--playlist-name can only be used with a single --xlsx file")

    if len(xlsx_paths) > 1 and args.output:
        raise PlaylistError("--output can only be used with a single --xlsx file; use --output-dir for multiple files")

    if args.output:
        output_dir = None
        output_file = Path(args.output).expanduser()
    else:
        output_dir = Path(args.output_dir).expanduser()
        output_file = None

    jobs: list[PlaylistJob] = []
    for xlsx in xlsx_paths:
        playlist_name = args.playlist_name or xlsx.stem
        output = output_file or output_dir / f"{playlist_name}.m3u"
        plex_path = None
        if args.plex_path:
            if len(xlsx_paths) > 1:
                raise PlaylistError("--plex-path can only be used with a single --xlsx file; use --plex-path-dir for multiple files")
            plex_path = args.plex_path
        elif args.plex_path_dir:
            plex_path = join_windows_path(args.plex_path_dir, f"{playlist_name}.m3u")
        jobs.append(PlaylistJob(xlsx=xlsx, playlist_name=playlist_name, output=output, plex_path=plex_path))

    return jobs


def join_windows_path(directory: str, filename: str) -> str:
    return directory.rstrip("\\/") + "\\" + filename


def validate_import_args(args: argparse.Namespace, token: str | None, jobs: list[PlaylistJob]) -> None:
    if args.generate_only and args.import_playlist:
        raise PlaylistError("--generate-only and --import cannot be used together")

    if not args.import_playlist:
        return

    if not args.plex_url:
        raise PlaylistError("--plex-url is required with --import")
    if not token:
        raise PlaylistError(
            "Plex token required. Pass --plex-token or set PLEX_TOKEN. "
            "Find it as X-Plex-Token in Plex Web URLs/XML, or PlexOnlineToken in PMS Preferences.xml."
        )
    if not args.section_id and not args.library_name:
        raise PlaylistError("--library-name or --section-id is required with --import")
    missing_plex_path = [job.playlist_name for job in jobs if not job.plex_path]
    if missing_plex_path:
        raise PlaylistError("--plex-path or --plex-path-dir is required with --import")


def run(args: argparse.Namespace) -> int:
    jobs = build_jobs(args)
    token = resolve_token(args)
    validate_import_args(args, token, jobs)

    section_id = args.section_id
    section_type = None
    if args.import_playlist and not section_id:
        assert token is not None
        section_id, section_type = find_library_section_id(args.plex_url, token, args.library_name, args.timeout)

    if args.dry_run:
        print("Mode: DRY RUN")
    elif args.generate_only:
        print("Mode: GENERATE ONLY")
    elif args.import_playlist:
        print("Mode: GENERATE AND IMPORT")
    else:
        print("Mode: GENERATE ONLY (default; pass --import to upload to Plex)")

    if section_id:
        library_suffix = f" ({section_type})" if section_type else ""
        print(f"Plex library section: {section_id}{library_suffix}")

    for job in jobs:
        parsed = parse_roon_excel(job.xlsx, args.sheet, args.path_column)
        playlist_paths = parsed.paths
        if args.path_mode == "relative":
            if not args.relative_root:
                raise PlaylistError("--relative-root is required with --path-mode relative")
            playlist_paths = to_relative_paths(playlist_paths, args.relative_root)

        print()
        print(f"Playlist: {job.playlist_name}")
        print(f"  Excel: {job.xlsx}")
        print(f"  Tracks read: {parsed.tracks_read}")
        print(f"  Paths written: {len(playlist_paths)}")
        print(f"  Blank paths skipped: {parsed.blank_paths}")
        print(f"  Output M3U: {job.output}")

        if args.dry_run:
            if args.import_playlist:
                print(f"  Plex import path: {job.plex_path}")
                print("  Import: skipped by dry run")
            continue

        write_m3u(job.output, playlist_paths, create_dirs=args.create_dirs)

        if args.import_playlist:
            assert token is not None
            assert section_id is not None
            assert job.plex_path is not None
            status, _body = upload_playlist(args.plex_url, token, section_id, job.plex_path, args.force, args.timeout)
            print(f"  Import: Plex HTTP {status}")
        else:
            print("  Import: skipped")

    return 0


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert Roon Excel playlist exports to M3U and optionally import them into Plex."
    )
    parser.add_argument("--xlsx", nargs="+", required=True, help="Roon Excel export(s) to convert")
    parser.add_argument("--playlist-name", help="Playlist name for a single Excel file; defaults to xlsx filename stem")
    parser.add_argument("--output", help="Output M3U file for a single Excel file")
    parser.add_argument("--output-dir", default=".", help="Output directory when --output is not used")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Excel sheet name")
    parser.add_argument("--path-column", default=DEFAULT_PATH_COLUMN, help="Excel column containing file paths")
    parser.add_argument("--path-mode", choices=["unc", "relative"], default="unc", help="Write UNC paths unchanged or relative paths")
    parser.add_argument("--relative-root", help="Root prefix to strip when using --path-mode relative")
    parser.add_argument("--create-dirs", action="store_true", help="Create output directories if missing")

    parser.add_argument("--plex-url", help="Plex Media Server URL, e.g. http://192.168.1.50:32400")
    parser.add_argument("--plex-token", help="Plex X-Plex-Token. Prefer PLEX_TOKEN env var to avoid shell history.")
    parser.add_argument("--library-name", default="Music", help="Plex library name to import into")
    parser.add_argument("--section-id", help="Plex library section ID; bypasses library-name lookup")
    parser.add_argument("--plex-path", help="PMS-visible path to output M3U for a single Excel file")
    parser.add_argument("--plex-path-dir", help="PMS-visible directory for generated M3U files")
    parser.add_argument("--force", action="store_true", help="Ask Plex to replace/update existing playlist when supported")
    parser.add_argument("--timeout", type=int, default=20, help="Plex request timeout in seconds")

    parser.add_argument("--dry-run", action="store_true", help="Validate and print planned actions without writing/importing")
    parser.add_argument("--generate-only", action="store_true", help="Create M3U file(s), skip Plex import")
    parser.add_argument("--import", dest="import_playlist", action="store_true", help="Import generated M3U file(s) into Plex")

    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        return run(args)
    except PlaylistError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
